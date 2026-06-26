#!/usr/bin/env python
import openpyxl
import re
import os
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.styles import Font, PatternFill
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText
from .estimators import *


def bold(text):
    return CellRichText(TextBlock(InlineFont(b=True), text))


def grey_text(cell, text):
    cell.value = text
    cell.font = Font(color='FFA9A9A9')


def yellow_cell(cell, text):
    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    cell.value = text


def create_MRT_estimation():
    wb = openpyxl.Workbook()

    ws = wb.active
    ws['A1'] = 'Time (s)'
    ws['B1'] = 'VO2 (l/min)'

    VO2_EX_2 = VO2_EX[12:]
    for i, val in enumerate(VO2_EX_2):
        ws[f'A{i + 2}'] = i * 10 + 120
        ws[f'B{i + 2}'] = val

    ws['D1'] = 'Final 90 s of baseline'
    VO2_WARMUP = VO2[WARMUP_SLICE][-9:]
    for i, val in enumerate(VO2_WARMUP):
        ws[f'D{i + 2}'] = val
        ws[f'E{i + 2}'] = val

    ws['D13'] = 'MRT'
    ws['E13'] = 'Calc baseline VO2 for last 60 s of baseline cycling'
    ws['D14'] = "MRT = (Baseline VO2 - 'C' from VO2 vs. Time scatter) / 'M' from same scatter"

    ws['D16'] = bold('MRT')
    ws['D17'] = bold('(s)')
    ws['D19'] = bold('GET')
    ws['D21'] = bold('(s)')
    ws['D22'] = bold('(W)')
    ws['C23'] = bold('Moderate')
    ws['D23'] = bold('90% GET')
    ws['D24'] = bold('30%Δ (mL.min)')
    ws['D25'] = bold('30%Δ (s)')
    ws['C26'] = bold('Heavy')
    ws['D26'] = bold('30%Δ (W)')
    ws['D27'] = bold('70%Δ (mL.min)')
    ws['D28'] = bold('70%Δ (s)')
    ws['C29'] = bold('Severe')
    ws['D29'] = bold('70%Δ (W)')
    ws['D30'] = bold('120%VO2max (mL.min)')
    ws['D31'] = bold('120%VO2max (s)')
    ws['C32'] = bold('Extreme')
    ws['D32'] = bold('120%VO2max (W)')
    ws['D33'] = bold('40%Δ (mL.min)')
    ws['D34'] = bold('40%Δ (s)')
    ws['C35'] = bold('Heavy')
    ws['D35'] = bold('40%Δ (W)')

    grey_text(ws['F16'], 'y')
    ws['G16'] = 'm'
    ws['H16'] = 'c'
    grey_text(ws['I16'], 'x')
    ws['J16'] = 'Baseline'

    ws['K17'] = 'final 60 s from baseline'
    yellow_cell(ws['J17'], '=AVERAGE(D5:D10)')
    slope_end = len(VO2_EX_2) - 20

    yellow_cell(ws['G17'], f'=SLOPE(B2:B{slope_end},A2:A{slope_end})')
    yellow_cell(ws['H17'], f'=INTERCEPT(B2:B{slope_end},A2:A{slope_end})')
    grey_text(ws['I17'], '=(J17-H17)/G17')
    grey_text(ws['E17'], ws['I17'].value)

    ws['F19'] = 'VO2 (l.min) from VE/VO2 plot'
    grey_text(ws['F20'], 'y')
    for i in range(20, 33, 3):
        grey_text(ws[f'G{i}'], 'm')
        grey_text(ws[f'G{i + 1}'], '=G17')
        grey_text(ws[f'H{i}'], 'c')
        grey_text(ws[f'H{i + 1}'], '=H17')
        grey_text(ws[f'I{i}'], 'x(s)')
        grey_text(ws[f'I{i + 1}'], f'=(E{i + 1}-H{i + 1})/G{i + 1}')
    grey_text(ws['I21'], '=(J21-H21)/G21')

    ws['J20'] = 'GET VO2'
    ws['K20'] = 'VO2max'
    yellow_cell(ws['K21'], VO2Max)
    ws['L20'] = 'Peak WR'
    yellow_cell(ws['L21'], max_watt)
    yellow_cell(ws['J21'], '')

    baseline_WR = watt[WARMUP_SLICE].min()
    ramp_WR = re.search(r'\d+(?=(\s)?[wW])', df.iloc[11, 4]).group(0)

    ws['M20'] = 'Ramp WR'
    yellow_cell(ws['M21'], ramp_WR)
    ws['N20'] = 'Baseline WR'
    yellow_cell(ws['N21'], baseline_WR)

    grey_text(ws['E21'], '=I21-E17')
    grey_text(ws['E22'], f'=((E21/60)*M21)+N21')
    ws['E23'] = '=E22*0.9'
    ws['E24'] = '=((K21-J21)*0.3)+J21'
    ws['E25'] = '=I24-E17'
    ws['E26'] = f'=((E25/60)*M21)+N21'
    ws['E27'] = '=((K21-J21)*0.7)+J21'
    ws['E28'] = '=I27-E17'
    ws['E29'] = f'=((E28/60)*M21)+N21'
    ws['E30'] = '=K21*1.2'
    ws['E31'] = '=I30-E17'
    ws['E32'] = f'=((E31/60)*M21)+N21'
    ws['E33'] = '=((K21-J21)*0.4)+J21'
    ws['E34'] = '=I33-E17'
    ws['E35'] = f'=((E34/60)*M21)+N21'

    # =========================
    # NEW: Additional targets
    # =========================
    ws['C37'] = bold('Additional')
    ws['D37'] = bold('Targets')

    # %GET targets (in Watts, consistent with your 90% GET approach)
    ws['D38'] = bold('80% GET (W)')
    ws['E38'] = '=E22*0.8'
    ws['D39'] = bold('60% GET (W)')
    ws['E39'] = '=E22*0.6'

    # Helper: VO2->time (s) using the same slope/intercept (G17/H17)
    # 25%Δ
    ws['D41'] = bold('25%Δ (mL.min)')
    ws['E41'] = '=((K21-J21)*0.25)+J21'
    grey_text(ws['I41'], '=(E41-$H$17)/$G$17')
    ws['D42'] = bold('25%Δ (s)')
    ws['E42'] = '=I41-E17'
    ws['D43'] = bold('25%Δ (W)')
    ws['E43'] = '=((E42/60)*M21)+N21'

    # 50%Δ
    ws['D45'] = bold('50%Δ (mL.min)')
    ws['E45'] = '=((K21-J21)*0.5)+J21'
    grey_text(ws['I45'], '=(E45-$H$17)/$G$17')
    ws['D46'] = bold('50%Δ (s)')
    ws['E46'] = '=I45-E17'
    ws['D47'] = bold('50%Δ (W)')
    ws['E47'] = '=((E46/60)*M21)+N21'

    # 60%Δ
    ws['D49'] = bold('60%Δ (mL.min)')
    ws['E49'] = '=((K21-J21)*0.6)+J21'
    grey_text(ws['I49'], '=(E49-$H$17)/$G$17')
    ws['D50'] = bold('60%Δ (s)')
    ws['E50'] = '=I49-E17'
    ws['D51'] = bold('60%Δ (W)')
    ws['E51'] = '=((E50/60)*M21)+N21'

    chart = openpyxl.chart.ScatterChart()
    chart.title = 'VO2 vs Time'
    chart.legend = None
    chart.x_axis.delete = False
    chart.y_axis.delete = False

    x_values = openpyxl.chart.Reference(ws, min_col=1, min_row=2, max_row=len(VO2_EX_2) + 1)
    y_values = openpyxl.chart.Reference(ws, min_col=2, min_row=2, max_row=len(VO2_EX_2) + 1)

    series = openpyxl.chart.Series(y_values, x_values, title_from_data=True)
    series.marker = openpyxl.chart.marker.Marker('circle', spPr=GraphicalProperties(solidFill='FF0000'))
    series.graphicalProperties.line.noFill = True
    chart.series.append(series)

    ws.add_chart(chart, 'L1')

    excel_file = f'{folder}/MRT estimation {participant_name}.xlsx'
    wb.save(excel_file)
    os.startfile(excel_file)
